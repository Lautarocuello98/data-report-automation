from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd


PALETTE = {
    "navy": "0F172A",
    "blue": "2563EB",
    "teal": "0F766E",
    "emerald": "15803D",
    "amber": "F59E0B",
    "orange": "EA580C",
    "slate": "475569",
    "sky": "E0F2FE",
    "mint": "D1FAE5",
    "sand": "FEF3C7",
    "rose": "FFEDD5",
    "mist": "F1F5F9",
    "white": "FFFFFF",
    "line": "CBD5E1",
}

THIN_BORDER = Border(
    left=Side(style="thin", color=PALETTE["line"]),
    right=Side(style="thin", color=PALETTE["line"]),
    top=Side(style="thin", color=PALETTE["line"]),
    bottom=Side(style="thin", color=PALETTE["line"]),
)


def _set_col_widths(ws, widths: dict[str, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _currency_token(currency: str) -> str:
    code = str(currency).strip().upper() or "USD"
    symbols = {
        "USD": "$",
        "EUR": "EUR",
        "GBP": "GBP",
        "JPY": "JPY",
        "ARS": "ARS",
        "MXN": "MXN",
        "GTQ": "GTQ",
    }
    return symbols.get(code, code)


def _currency_number_format(currency: str) -> str:
    token = _currency_token(currency)
    return f'"{token}" #,##0.00;[Red]-"{token}" #,##0.00'


def _currency_text(value: float, currency: str) -> str:
    token = _currency_token(currency)
    if token == "$":
        return f"{token}{value:,.2f}"
    return f"{token} {value:,.2f}"


def _apply_number_format_to_column(ws, col_idx: int, number_format: str, start_row: int = 2) -> None:
    for row_idx in range(start_row, ws.max_row + 1):
        ws.cell(row=row_idx, column=col_idx).number_format = number_format


def _header_positions(ws, header_row: int = 1) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col_idx).value
        if value is not None:
            headers[str(value).strip()] = col_idx
    return headers


def _fill_range(
    ws,
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int,
    *,
    fill: PatternFill | None = None,
    font: Font | None = None,
    alignment: Alignment | None = None,
    border: Border | None = None,
) -> None:
    for row in ws.iter_rows(
        min_row=start_row,
        max_row=end_row,
        min_col=start_col,
        max_col=end_col,
    ):
        for cell in row:
            if fill is not None:
                cell.fill = fill
            if font is not None:
                cell.font = font
            if alignment is not None:
                cell.alignment = alignment
            if border is not None:
                cell.border = border


def _merge_block(
    ws,
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int,
    value,
    *,
    fill_color: str,
    font: Font,
    alignment: Alignment,
    border: Border = THIN_BORDER,
    number_format: str | None = None,
) -> None:
    ws.merge_cells(
        start_row=start_row,
        start_column=start_col,
        end_row=end_row,
        end_column=end_col,
    )
    _fill_range(
        ws,
        start_row,
        end_row,
        start_col,
        end_col,
        fill=PatternFill("solid", fgColor=fill_color),
        font=font,
        alignment=alignment,
        border=border,
    )
    cell = ws.cell(row=start_row, column=start_col)
    cell.value = value
    if number_format is not None:
        cell.number_format = number_format


def _auto_fit_widths(
    ws,
    *,
    min_width: int = 12,
    max_width: int = 28,
    extra: int = 2,
    max_col: int | None = None,
) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if max_col is not None and cell.column > max_col:
                continue
            if cell.value is None:
                continue
            text_length = len(str(cell.value))
            widths[cell.column] = max(widths.get(cell.column, min_width), min(text_length + extra, max_width))
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _style_table(
    ws,
    *,
    header_row: int = 1,
    end_col: int | None = None,
    end_row: int | None = None,
) -> None:
    table_end_col = end_col or ws.max_column
    table_end_row = end_row or ws.max_row
    if table_end_row < header_row or table_end_col <= 0:
        return

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(table_end_col)}{table_end_row}"
    ws.row_dimensions[header_row].height = 24

    for col_idx in range(1, table_end_col + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = PatternFill("solid", fgColor=PALETTE["navy"])
        cell.font = Font(color=PALETTE["white"], bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    band_a = PatternFill("solid", fgColor=PALETTE["white"])
    band_b = PatternFill("solid", fgColor=PALETTE["mist"])
    for row_idx in range(header_row + 1, table_end_row + 1):
        fill = band_a if row_idx % 2 == 0 else band_b
        for col_idx in range(1, table_end_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.border = THIN_BORDER
            if cell.alignment == Alignment():
                cell.alignment = Alignment(vertical="center")


def _format_date_label(value: str | None) -> str:
    if not value:
        return "N/A"
    try:
        return pd.to_datetime(value).strftime("%b %d, %Y")
    except Exception:
        return str(value)


def _write_dataframe_rows(ws, df: pd.DataFrame, *, start_row: int) -> None:
    for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)


def _top_group_label(count: int, *, title_case: bool = False) -> str:
    if count <= 0:
        return "products"
    prefix = "top"
    noun = "product" if count == 1 else "products"
    label = f"{prefix} {count} {noun}"
    return label.title() if title_case else label


def _top_group_verb(count: int) -> str:
    return "contributes" if count == 1 else "contribute"


def _counted_products(count: int) -> str:
    noun = "product" if count == 1 else "products"
    return f"{count} {noun}"


def _margin_position(margin_pct: float) -> str:
    if margin_pct < 0:
        return "loss-making economics"
    if margin_pct >= 35:
        return "strong profitability"
    if margin_pct >= 20:
        return "healthy profitability"
    if margin_pct > 0:
        return "positive but thin profitability"
    return "break-even performance"


def _margin_signal(margin_pct: float) -> str:
    if margin_pct < 0:
        return "Margin is negative"
    if margin_pct >= 35:
        return "Margin is strong"
    if margin_pct >= 20:
        return "Margin is healthy"
    if margin_pct > 0:
        return "Margin is thin"
    return "Margin is at break-even"


def _concentration_profile(kpis: dict) -> dict[str, int | float | str]:
    available_products = int(kpis.get("ranked_products_count", 0) or 0)
    top_count = min(3, available_products) if available_products else 0
    top_product_share = float(kpis.get("top_product_share", 0.0))
    top_3_share = float(kpis.get("top_3_products_share", 0.0))

    if available_products <= 0:
        level = "unknown"
    elif available_products <= 3:
        level = "narrow"
    elif top_3_share >= 75 or top_product_share >= 45:
        level = "high"
    elif top_3_share >= 55 or top_product_share >= 30:
        level = "medium"
    else:
        level = "balanced"

    return {
        "available_products": available_products,
        "top_count": top_count,
        "top_product_share": top_product_share,
        "top_3_share": top_3_share,
        "level": level,
    }


def _concentration_clause(kpis: dict) -> str:
    profile = _concentration_profile(kpis)
    available_products = int(profile["available_products"])
    top_count = int(profile["top_count"])
    top_3_share = float(profile["top_3_share"])
    level = str(profile["level"])

    if level == "unknown":
        return "product concentration could not be assessed from the available data"
    if level == "narrow":
        return f"only {_counted_products(available_products)} are represented, so the mix is naturally concentrated"
    if level == "high":
        return (
            f"{_top_group_label(top_count)} {_top_group_verb(top_count)} {top_3_share:.1f}% "
            "of revenue, indicating high dependence on a few products"
        )
    if level == "medium":
        return (
            f"{_top_group_label(top_count)} {_top_group_verb(top_count)} {top_3_share:.1f}% "
            "of revenue, so a few products are carrying most sales"
        )
    return (
        f"{_top_group_label(top_count)} {_top_group_verb(top_count)} {top_3_share:.1f}% "
        "of revenue, suggesting a balanced product mix"
    )


def _build_executive_summary_text(kpis: dict, currency: str) -> str:
    total_revenue = float(kpis.get("total_revenue", 0.0))
    total_profit = float(kpis.get("total_profit", 0.0))
    total_orders = int(kpis.get("total_orders", 0))
    margin_pct = float(kpis.get("margin_pct", 0.0))
    best_product = kpis.get("best_product") or {}
    best_day = kpis.get("best_day") or {}
    top_product_name = str(best_product.get("product", "N/A"))
    top_product_revenue = float(best_product.get("revenue", 0.0))
    top_product_share = float(kpis.get("top_product_share", 0.0))
    peak_day_share = float(kpis.get("peak_day_share", 0.0))

    if total_orders == 0 and total_revenue == 0:
        return (
            "No commercial activity is available in the current dataset, so revenue, profitability, "
            "and peak-day signals are limited."
        )

    summary_lines = [
        (
            f"Revenue reached {_currency_text(total_revenue, currency)} across {total_orders:,} orders, "
            f"generating {_currency_text(total_profit, currency)} in profit at a {margin_pct:.1f}% margin, "
            f"which points to {_margin_position(margin_pct)}."
        )
    ]

    if top_product_revenue > 0:
        summary_lines.append(
            f"{top_product_name} led the catalog with {_currency_text(top_product_revenue, currency)} "
            f"({top_product_share:.1f}% of revenue); {_concentration_clause(kpis)}."
        )
    else:
        summary_lines.append(f"{_concentration_clause(kpis).capitalize()}.")

    if best_day:
        summary_lines.append(
            f"Peak day was {_format_date_label(best_day.get('date'))} with "
            f"{_currency_text(float(best_day.get('revenue', 0.0)), currency)} in revenue "
            f"({peak_day_share:.1f}% of the period total)."
        )
    else:
        summary_lines.append(
            "Peak-day performance could not be assessed because the dataset does not contain valid dates."
        )

    return " ".join(summary_lines)


def _build_summary_insight_text(kpis: dict) -> str:
    total_orders = int(kpis.get("total_orders", 0))
    total_revenue = float(kpis.get("total_revenue", 0.0))
    margin_pct = float(kpis.get("margin_pct", 0.0))
    profile = _concentration_profile(kpis)
    top_count = int(profile["top_count"])
    top_3_share = float(profile["top_3_share"])
    available_products = int(profile["available_products"])
    level = str(profile["level"])

    if total_orders == 0 and total_revenue == 0:
        return "The dataset is too limited to support a business conclusion."

    margin_part = _margin_signal(margin_pct)
    if level == "unknown":
        return f"{margin_part}, and product concentration cannot be assessed from the available detail."
    if level == "narrow":
        return f"{margin_part}, but the mix is narrow because only {_counted_products(available_products)} are represented."
    if level == "high":
        return (
            f"{margin_part}, but revenue is highly concentrated: "
            f"{_top_group_label(top_count)} {_top_group_verb(top_count)} {top_3_share:.1f}% of sales."
        )
    if level == "medium":
        return (
            f"{margin_part}, and a few products are doing most of the work: "
            f"{_top_group_label(top_count)} {_top_group_verb(top_count)} {top_3_share:.1f}% of sales."
        )
    return (
        f"{margin_part}, and revenue is reasonably diversified: "
        f"{_top_group_label(top_count)} {_top_group_verb(top_count)} {top_3_share:.1f}% of sales."
    )


def _build_top_products_insight_text(kpis: dict) -> str:
    profile = _concentration_profile(kpis)
    top_count = int(profile["top_count"])
    top_3_share = float(profile["top_3_share"])
    available_products = int(profile["available_products"])
    level = str(profile["level"])

    if level == "unknown":
        return "Not enough product-level revenue is available to assess concentration."
    if level == "narrow":
        return f"Only {_counted_products(available_products)} appear in the dataset, so the sales mix is naturally concentrated."
    if level == "high":
        return (
            f"{_top_group_label(top_count)} {_top_group_verb(top_count)} {top_3_share:.1f}% of revenue, "
            "so a small set of products is carrying the business."
        )
    if level == "medium":
        return (
            f"{_top_group_label(top_count)} {_top_group_verb(top_count)} {top_3_share:.1f}% of revenue, "
            "showing moderate concentration in the catalog."
        )
    return (
        f"{_top_group_label(top_count)} {_top_group_verb(top_count)} {top_3_share:.1f}% of revenue, "
        "suggesting a balanced product mix."
    )


def _write_dataframe_sheet(
    wb: Workbook,
    title: str,
    df: pd.DataFrame,
    *,
    currency_fmt: str,
    date_cols: Iterable[str] = (),
    currency_cols: Iterable[str] = (),
    percent_cols: Iterable[str] = (),
    decimal_cols: Iterable[str] = (),
    integer_cols: Iterable[str] = (),
) -> None:
    ws = wb.create_sheet(title)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    headers = _header_positions(ws)
    for col_name in date_cols:
        if col_name in headers:
            _apply_number_format_to_column(ws, headers[col_name], "yyyy-mm-dd")
    for col_name in currency_cols:
        if col_name in headers:
            _apply_number_format_to_column(ws, headers[col_name], currency_fmt)
    for col_name in percent_cols:
        if col_name in headers:
            _apply_number_format_to_column(ws, headers[col_name], '0.0"%"')
    for col_name in decimal_cols:
        if col_name in headers:
            _apply_number_format_to_column(ws, headers[col_name], "#,##0.00")
    for col_name in integer_cols:
        if col_name in headers:
            _apply_number_format_to_column(ws, headers[col_name], "#,##0")

    _style_table(ws)
    _auto_fit_widths(ws)


def _write_kpi_card(
    ws,
    *,
    start_row: int,
    start_col: int,
    end_col: int,
    title: str,
    value,
    note: str,
    fill_color: str,
    number_format: str | None = None,
) -> None:
    _merge_block(
        ws,
        start_row,
        start_row,
        start_col,
        end_col,
        title,
        fill_color=fill_color,
        font=Font(color=PALETTE["white"], bold=True, size=11),
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    _merge_block(
        ws,
        start_row + 1,
        start_row + 2,
        start_col,
        end_col,
        value,
        fill_color=fill_color,
        font=Font(color=PALETTE["white"], bold=True, size=24),
        alignment=Alignment(horizontal="center", vertical="center"),
        number_format=number_format,
    )
    _merge_block(
        ws,
        start_row + 3,
        start_row + 3,
        start_col,
        end_col,
        note,
        fill_color=fill_color,
        font=Font(color=PALETTE["white"], italic=True, size=10),
        alignment=Alignment(horizontal="center", vertical="center"),
    )


def _write_insight_tile(
    ws,
    *,
    start_row: int,
    start_col: int,
    end_col: int,
    title: str,
    value: str,
    note: str,
    fill_color: str,
) -> None:
    _merge_block(
        ws,
        start_row,
        start_row,
        start_col,
        end_col,
        title,
        fill_color=fill_color,
        font=Font(color=PALETTE["navy"], bold=True, size=10),
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    _merge_block(
        ws,
        start_row + 1,
        start_row + 2,
        start_col,
        end_col,
        value,
        fill_color=fill_color,
        font=Font(color=PALETTE["navy"], bold=True, size=16),
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
    )
    _merge_block(
        ws,
        start_row + 3,
        start_row + 3,
        start_col,
        end_col,
        note,
        fill_color=fill_color,
        font=Font(color=PALETTE["slate"], italic=True, size=9),
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
    )


def _embed_chart(ws, image_path: Path, anchor: str, *, width: int, height: int) -> None:
    try:
        image = XLImage(str(image_path))
        image.width = width
        image.height = height
        image.anchor = anchor
        ws.add_image(image)
    except Exception as exc:
        logging.warning("Could not embed chart '%s': %s", image_path, exc)


def _build_summary_sheet(
    ws,
    *,
    kpis: dict,
    currency: str,
    sources: list[str],
    chart_files: list[Path] | None,
) -> None:
    currency_fmt = _currency_number_format(currency)
    total_revenue = float(kpis.get("total_revenue", 0.0))
    total_profit = float(kpis.get("total_profit", 0.0))
    total_cost = float(kpis.get("total_cost", 0.0))
    total_orders = int(kpis.get("total_orders", 0))
    total_units = float(kpis.get("total_units", 0.0))
    avg_order_value = float(kpis.get("avg_order_value", 0.0))
    margin_pct = float(kpis.get("margin_pct", 0.0))
    unique_products = int(kpis.get("unique_products", 0))
    unique_skus = int(kpis.get("unique_skus", 0))
    best_product = kpis.get("best_product") or {}
    best_day = kpis.get("best_day") or {}
    coverage = (
        f"{_format_date_label(kpis.get('date_start'))} to {_format_date_label(kpis.get('date_end'))}"
        if kpis.get("date_start") and kpis.get("date_end")
        else "Coverage unavailable"
    )
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    top_product_name = str(best_product.get("product", "N/A"))
    top_product_revenue = float(best_product.get("revenue", 0.0))
    top_product_share = float(kpis.get("top_product_share", 0.0))
    summary_insight = _build_summary_insight_text(kpis)
    peak_day_note = (
        f"{_currency_text(float(best_day.get('revenue', 0.0)), currency)} generated"
        if best_day
        else "No valid dates available"
    )
    revenue_leader_note = (
        f"{_currency_text(top_product_revenue, currency)} in revenue"
        if top_product_revenue > 0
        else "No product revenue available"
    )

    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = PALETTE["blue"]
    _set_col_widths(
        ws,
        {
            "A": 14,
            "B": 14,
            "C": 14,
            "D": 14,
            "E": 14,
            "F": 14,
            "G": 14,
            "H": 14,
            "I": 14,
            "J": 14,
            "K": 14,
            "L": 14,
        },
    )

    _fill_range(
        ws,
        1,
        60,
        1,
        12,
        fill=PatternFill("solid", fgColor=PALETTE["mist"]),
    )

    _merge_block(
        ws,
        1,
        2,
        1,
        12,
        f"Sales Performance Dashboard ({currency})",
        fill_color=PALETTE["navy"],
        font=Font(color=PALETTE["white"], bold=True, size=22),
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    _merge_block(
        ws,
        3,
        3,
        1,
        12,
        f"Generated on {generated_at} | {len(sources)} source file(s) | {coverage}",
        fill_color=PALETTE["slate"],
        font=Font(color=PALETTE["white"], size=10),
        alignment=Alignment(horizontal="center", vertical="center"),
    )

    _write_kpi_card(
        ws,
        start_row=5,
        start_col=1,
        end_col=4,
        title="Total Orders",
        value=total_orders,
        note="Commercial transactions processed",
        fill_color=PALETTE["blue"],
        number_format="#,##0",
    )
    _write_kpi_card(
        ws,
        start_row=5,
        start_col=5,
        end_col=8,
        title="Total Revenue",
        value=total_revenue,
        note="Gross sales generated in the period",
        fill_color=PALETTE["navy"],
        number_format=currency_fmt,
    )
    _write_kpi_card(
        ws,
        start_row=5,
        start_col=9,
        end_col=12,
        title="Total Profit",
        value=total_profit,
        note="Bottom-line contribution after cost",
        fill_color=PALETTE["emerald"],
        number_format=currency_fmt,
    )
    _write_kpi_card(
        ws,
        start_row=10,
        start_col=1,
        end_col=4,
        title="Avg Order Value",
        value=avg_order_value,
        note="Average ticket per order",
        fill_color=PALETTE["amber"],
        number_format=currency_fmt,
    )
    _write_kpi_card(
        ws,
        start_row=10,
        start_col=5,
        end_col=8,
        title="Active Products",
        value=f"{unique_products:,}",
        note=f"{unique_skus:,} SKUs represented in the dataset",
        fill_color=PALETTE["teal"],
    )
    _write_kpi_card(
        ws,
        start_row=10,
        start_col=9,
        end_col=12,
        title="Margin Rate",
        value=margin_pct,
        note=f"Total cost base: {_currency_text(total_cost, currency)}",
        fill_color=PALETTE["orange"],
        number_format='0.0"%"',
    )

    _merge_block(
        ws,
        15,
        15,
        1,
        12,
        "Executive Summary",
        fill_color=PALETTE["navy"],
        font=Font(color=PALETTE["white"], bold=True, size=12),
        alignment=Alignment(horizontal="left", vertical="center"),
    )
    executive_summary = (
        _build_executive_summary_text(kpis, currency)
    )
    _merge_block(
        ws,
        16,
        18,
        1,
        12,
        executive_summary,
        fill_color=PALETTE["white"],
        font=Font(color=PALETTE["navy"], size=11),
        alignment=Alignment(horizontal="left", vertical="top", wrap_text=True),
    )
    _merge_block(
        ws,
        20,
        20,
        1,
        12,
        f"So What? {summary_insight}",
        fill_color=PALETTE["sky"],
        font=Font(color=PALETTE["navy"], bold=True, italic=True, size=10),
        alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
    )
    ws.row_dimensions[19].height = 8

    _write_insight_tile(
        ws,
        start_row=21,
        start_col=1,
        end_col=3,
        title="Revenue Leader",
        value=top_product_name,
        note=revenue_leader_note,
        fill_color=PALETTE["sky"],
    )
    _write_insight_tile(
        ws,
        start_row=21,
        start_col=4,
        end_col=6,
        title="Peak Day",
        value=_format_date_label(best_day.get("date")),
        note=peak_day_note,
        fill_color=PALETTE["mint"],
    )
    _write_insight_tile(
        ws,
        start_row=21,
        start_col=7,
        end_col=9,
        title="Revenue Share",
        value=f"{top_product_share:.1f}%",
        note="Contribution from the leading product",
        fill_color=PALETTE["sand"],
    )
    _write_insight_tile(
        ws,
        start_row=21,
        start_col=10,
        end_col=12,
        title="Catalog Spread",
        value=f"{unique_products:,} products",
        note=f"{unique_skus:,} SKUs active in the report",
        fill_color=PALETTE["rose"],
    )

    _merge_block(
        ws,
        25,
        25,
        1,
        12,
        "Visual Story",
        fill_color=PALETTE["navy"],
        font=Font(color=PALETTE["white"], bold=True, size=12),
        alignment=Alignment(horizontal="left", vertical="center"),
    )
    if chart_files:
        for image_path, anchor in zip(chart_files[:2], ["A27", "G27"], strict=False):
            _embed_chart(ws, image_path, anchor, width=520, height=280)

    source_start_row = 47
    _merge_block(
        ws,
        source_start_row,
        source_start_row,
        1,
        12,
        "Source Files",
        fill_color=PALETTE["navy"],
        font=Font(color=PALETTE["white"], bold=True, size=12),
        alignment=Alignment(horizontal="left", vertical="center"),
    )
    for idx, source in enumerate(sources, start=source_start_row + 1):
        _merge_block(
            ws,
            idx,
            idx,
            1,
            12,
            source,
            fill_color=PALETTE["white"],
            font=Font(color=PALETTE["navy"], size=10),
            alignment=Alignment(horizontal="left", vertical="center"),
        )


def _build_top_products_sheet(
    wb: Workbook,
    *,
    product_performance: pd.DataFrame,
    kpis: dict,
    currency: str,
) -> None:
    ws = wb.create_sheet("Top Products")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = PALETTE["amber"]
    _set_col_widths(ws, {get_column_letter(col): 14 for col in range(1, 13)})
    _fill_range(
        ws,
        1,
        40,
        1,
        12,
        fill=PatternFill("solid", fgColor=PALETTE["mist"]),
    )

    top_count = min(int(kpis.get("ranked_products_count", 0) or 0), 3)
    top_group_title = _top_group_label(top_count, title_case=True) if top_count else "Top Products"
    top_3_revenue = float(kpis.get("top_3_products_revenue", 0.0))
    top_3_share = float(kpis.get("top_3_products_share", 0.0))
    best_product = kpis.get("best_product") or {}
    top_product_name = str(best_product.get("product", "N/A"))
    top_product_revenue = float(best_product.get("revenue", 0.0))
    top_product_share = float(kpis.get("top_product_share", 0.0))

    _merge_block(
        ws,
        1,
        2,
        1,
        12,
        "Top Product Performance",
        fill_color=PALETTE["navy"],
        font=Font(color=PALETTE["white"], bold=True, size=20),
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    _merge_block(
        ws,
        3,
        3,
        1,
        12,
        "Revenue contribution and concentration across the leading products.",
        fill_color=PALETTE["slate"],
        font=Font(color=PALETTE["white"], size=10),
        alignment=Alignment(horizontal="center", vertical="center"),
    )

    _write_kpi_card(
        ws,
        start_row=5,
        start_col=1,
        end_col=4,
        title=f"{top_group_title} Revenue",
        value=top_3_revenue,
        note="Combined revenue from the leading products",
        fill_color=PALETTE["navy"],
        number_format=_currency_number_format(currency),
    )
    _write_kpi_card(
        ws,
        start_row=5,
        start_col=5,
        end_col=8,
        title=f"{top_group_title} Share",
        value=top_3_share,
        note="Share of total revenue",
        fill_color=PALETTE["blue"],
        number_format='0.0"%"',
    )
    _write_insight_tile(
        ws,
        start_row=5,
        start_col=9,
        end_col=12,
        title="Revenue Leader",
        value=top_product_name,
        note=(
            f"{_currency_text(top_product_revenue, currency)} | {top_product_share:.1f}% of sales"
            if top_product_revenue > 0
            else "No leading product identified"
        ),
        fill_color=PALETTE["mint"],
    )

    _merge_block(
        ws,
        10,
        10,
        1,
        12,
        "Product Concentration",
        fill_color=PALETTE["navy"],
        font=Font(color=PALETTE["white"], bold=True, size=12),
        alignment=Alignment(horizontal="left", vertical="center"),
    )
    _merge_block(
        ws,
        11,
        12,
        1,
        12,
        _build_top_products_insight_text(kpis),
        fill_color=PALETTE["white"],
        font=Font(color=PALETTE["navy"], size=11),
        alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
    )

    top_products_df = product_performance.head(10).copy()
    if top_products_df.empty:
        _merge_block(
            ws,
            14,
            16,
            1,
            12,
            "No product-level sales rows are available for ranking.",
            fill_color=PALETTE["white"],
            font=Font(color=PALETTE["navy"], size=11),
            alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
        )
        return

    table_start_row = 14
    table_end_col = len(top_products_df.columns)
    table_end_row = table_start_row + len(top_products_df)
    _write_dataframe_rows(ws, top_products_df, start_row=table_start_row)
    headers = _header_positions(ws, header_row=table_start_row)
    for col_name in ["revenue", "cost", "profit"]:
        if col_name in headers:
            _apply_number_format_to_column(ws, headers[col_name], _currency_number_format(currency), start_row=table_start_row + 1)
    if "margin_pct" in headers:
        _apply_number_format_to_column(ws, headers["margin_pct"], '0.0"%"', start_row=table_start_row + 1)
    if "total_units" in headers:
        _apply_number_format_to_column(ws, headers["total_units"], "#,##0.00", start_row=table_start_row + 1)
    if "order_count" in headers:
        _apply_number_format_to_column(ws, headers["order_count"], "#,##0", start_row=table_start_row + 1)

    _style_table(ws, header_row=table_start_row, end_col=table_end_col, end_row=table_end_row)
    _auto_fit_widths(ws, max_col=table_end_col)


def _build_charts_sheet(wb: Workbook, chart_files: list[Path]) -> None:
    if not chart_files:
        return

    ws = wb.create_sheet("Charts")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = PALETTE["teal"]
    _set_col_widths(ws, {get_column_letter(col): 14 for col in range(1, 15)})
    _fill_range(
        ws,
        1,
        50,
        1,
        14,
        fill=PatternFill("solid", fgColor=PALETTE["mist"]),
    )
    _merge_block(
        ws,
        1,
        2,
        1,
        14,
        "Sales Performance Breakdown",
        fill_color=PALETTE["navy"],
        font=Font(color=PALETTE["white"], bold=True, size=20),
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    _merge_block(
        ws,
        3,
        3,
        1,
        14,
        "A clean set of visuals for fast commercial review.",
        fill_color=PALETTE["slate"],
        font=Font(color=PALETTE["white"], size=10),
        alignment=Alignment(horizontal="center", vertical="center"),
    )

    anchors = ["A6", "H6", "A25", "H25"]
    for image_path, anchor in zip(chart_files[:4], anchors, strict=False):
        _embed_chart(ws, image_path, anchor, width=600, height=320)


def generate_excel_report(
    report_path: Path,
    df_clean: pd.DataFrame,
    kpis: dict,
    currency: str,
    sources: Iterable[str],
    chart_files: list[Path] | None = None,
) -> None:
    """
    Creates an Excel report with:
    - Summary dashboard
    - Cleaned Data sheet
    - Product analysis sheets
    - Optional chart gallery
    """
    report_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    currency_fmt = _currency_number_format(currency)
    source_list = list(sources)

    summary_ws = wb.active
    _build_summary_sheet(
        summary_ws,
        kpis=kpis,
        currency=currency,
        sources=source_list,
        chart_files=chart_files,
    )

    df_out = kpis.get("df_with_calculations")
    if not isinstance(df_out, pd.DataFrame):
        df_out = df_clean.copy()
        if {"quantity", "unit_price", "unit_cost"}.issubset(df_out.columns):
            if "revenue" not in df_out.columns:
                df_out["revenue"] = df_out["quantity"] * df_out["unit_price"]
            if "cost" not in df_out.columns:
                df_out["cost"] = df_out["quantity"] * df_out["unit_cost"]
            if "profit" not in df_out.columns:
                df_out["profit"] = df_out["revenue"] - df_out["cost"]

    _write_dataframe_sheet(
        wb,
        "Cleaned Data",
        df_out,
        currency_fmt=currency_fmt,
        date_cols=["date"],
        currency_cols=["unit_price", "unit_cost", "revenue", "cost", "profit"],
        percent_cols=["margin_pct"],
        decimal_cols=["quantity"],
    )

    product_performance = kpis.get("product_performance")
    if not isinstance(product_performance, pd.DataFrame):
        product_performance = pd.DataFrame(
            columns=["product", "order_count", "total_units", "revenue", "cost", "profit", "margin_pct"]
        )
    _build_top_products_sheet(
        wb,
        product_performance=product_performance,
        kpis=kpis,
        currency=currency,
    )
    if not product_performance.empty:
        _write_dataframe_sheet(
            wb,
            "Product Performance",
            product_performance,
            currency_fmt=currency_fmt,
            currency_cols=["revenue", "cost", "profit"],
            percent_cols=["margin_pct"],
            decimal_cols=["total_units"],
            integer_cols=["order_count"],
        )

    daily_performance = kpis.get("daily_performance")
    if isinstance(daily_performance, pd.DataFrame) and not daily_performance.empty:
        _write_dataframe_sheet(
            wb,
            "Daily Performance",
            daily_performance,
            currency_fmt=currency_fmt,
            date_cols=["date"],
            currency_cols=["revenue", "cost", "profit"],
            percent_cols=["margin_pct"],
            decimal_cols=["units"],
            integer_cols=["orders"],
        )

    if chart_files:
        _build_charts_sheet(wb, chart_files)

    wb.save(report_path)
