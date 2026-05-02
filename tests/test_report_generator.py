from pathlib import Path

from openpyxl import load_workbook

from src.charts import build_charts
from src.processor import compute_kpis
from src.report_generator import generate_excel_report


def test_generate_excel_report_creates_file(tmp_path: Path, sample_df):
    kpis = compute_kpis(sample_df)
    out = tmp_path / "out.xlsx"

    generate_excel_report(
        report_path=out,
        df_clean=sample_df,
        kpis=kpis,
        currency="USD",
        sources=["sample.csv"],
        chart_files=[],
    )

    assert out.exists()
    assert out.stat().st_size > 0


def test_generate_excel_report_formats_currency_cells(tmp_path: Path, sample_df):
    kpis = compute_kpis(sample_df)
    out = tmp_path / "out.xlsx"
    chart_files = build_charts(sample_df, kpis, tmp_path / "charts")

    generate_excel_report(
        report_path=out,
        df_clean=sample_df,
        kpis=kpis,
        currency="USD",
        sources=["sample.csv"],
        chart_files=chart_files,
    )

    wb = load_workbook(out)
    summary = wb["Summary"]
    money_fmt = summary["E6"].number_format  # Total Revenue card value

    assert str(summary["A1"].value).startswith("Sales Performance Dashboard")
    assert "Revenue reached $40.00 across 2 orders" in str(summary["A16"].value)
    assert "Widget led the catalog with $20.00" in str(summary["A16"].value)
    assert "Peak day was Jan" in str(summary["A16"].value)
    assert "$20.00 in revenue (50.0% of the period total)" in str(summary["A16"].value)
    assert summary["A19"].value is None
    assert str(summary["A20"].value).startswith("So What?")
    assert summary["A20"].font.italic is True
    assert "Margin is strong" in str(summary["A20"].value)
    assert "only 2 products are represented" in str(summary["A20"].value)
    assert "$" in money_fmt
    assert "#,##0.00" in money_fmt
    assert "Charts" in wb.sheetnames
    assert "Product Performance" in wb.sheetnames
    assert "Daily Performance" in wb.sheetnames

    cleaned = wb["Cleaned Data"]
    cleaned_headers = {
        cleaned.cell(row=1, column=idx).value: idx
        for idx in range(1, cleaned.max_column + 1)
    }
    unit_price_col = cleaned_headers["unit_price"]
    assert cleaned.cell(row=2, column=unit_price_col).number_format == money_fmt

    top = wb["Top Products"]
    assert top["A1"].value == "Top Product Performance"
    assert top["A5"].value == "Top 2 Products Revenue"
    assert top["A6"].value == 40.0
    assert top["E5"].value == "Top 2 Products Share"
    assert top["E6"].value == 100.0
    assert "Only 2 products appear in the dataset" in str(top["A11"].value)
    assert top["A14"].value == "product"
    assert top.auto_filter.ref == "A14:G16"
    revenue_col = {top.cell(row=14, column=idx).value: idx for idx in range(1, top.max_column + 1)}["revenue"]
    assert top.cell(row=15, column=revenue_col).number_format == money_fmt

    charts = wb["Charts"]
    assert charts["A1"].value == "Sales Performance Breakdown"
